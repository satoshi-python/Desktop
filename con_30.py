
import tool
import numpy as np
import MDAnalysis
from argparse import ArgumentParser
import openpyxl
import threading

def get_option():
    argparser = ArgumentParser()
    argparser.add_argument("-trr","--trajectory",type=str,help="path of trr")
    argparser.add_argument("-pdb","--protein",type=str,help="path of pdb")
    argparser.add_argument("-prob","--probtxt",type=str,help="path of pdb")
    argparser.add_argument("-name","--name_kei",type=str,help="name of file")
    return argparser.parse_args()

class con_thle(threading.Thread):
    def __init__(self, thread_name, frm, MED, PROB):
        self.frm = frm
        self.prob = PROB
        self.MED = MED
        self.MED26 = [0.0 for i in range(92)]
        self.thread_name = thread_name
        threading.Thread.__init__(self)

    def __str__(self):
        return self.thread_name

    def CAL_LI(self, cor, num):
        num1 = 0
        global MED
        for j in range(92,114):
            for j1 in self.MED[j]:
                a = tool.contact_list(cor[num],cor[j1])
                if a <= 8:
                    num1 += 1
                    break
                elif a >=15:
                    break
            if num1 == 1:
                break
        return num1

    def CAL_COR_MED(self, cor):
        kai = []
        for i in range(92):
            for j1 in self.MED[i]:
                jj = self.CAL_LI(cor,j1)
                if jj == 1:
                    break
            kai.append(int(jj))
        return kai

    def run(self):
        print("start:",self.thread_name,"len TRR:",len(self.frm),"LEN PROB:",len(self.prob))
        for i in range(len(self.frm)):
            if self.prob[i] > 0:
                cor = self.frm[i]
                aa = self.CAL_COR_MED(cor)
                for jj in range(len(self.MED26)):
                    if aa[jj] == 1:
                        self.MED26[jj] += float(self.prob[i])
                # if i % 100 == 0 and i != 0:
                    # print(self.thread_name,":",i)
        # book = openpyxl.load_workbook("test_20.xlsx")
        # sheet = book[str(self.NAME)]
        # for jj in range(92):
            # sheet.cell(row=2+jj,column=(int(self.thread_name))).value = float(self.MED26[jj])
        # book.save('test_20.xlsx')
        print("end:", self.thread_name)

    def get_value(self):
        return self.MED26

def main():
    thread_list = []
    args = get_option()
    book = openpyxl.load_workbook("test_30.xlsx")
    book.create_sheet(str(args.name_kei))
    book.save("test_30.xlsx")
    num1 = 1
    num2 = 0
    kai = []
    MED = []
    for i in open(str(args.protein)):
        f = i.split()
        if int(f[5]) == num1:
            kai.append(int(f[1]))
        else:
            MED.append(kai)
            kai=[int(f[1])]
            num1 += 1
            if f[4] == "B" and num2 == 0:
                num2 += 1
                num1 = 1
    MED.append(kai)
    PROB = []
    for i in open(str(args.probtxt)):
        PROB.append(float(i))
    u = MDAnalysis.Universe(str(args.trajectory))
    frm = u.trajectory
    kai1 = 0.0
    for i in range(1,181):
        thread = con_thle(i, frm[(i-1) * 5000:(i-1) * 5000 + 5000], MED, PROB[(i-1) * 5000:(i-1) * 5000 + 5000])
        print((i-1) * 5000,":",(i-1) * 5000 + 5000)
        thread.start()
        thread_list.append(thread)
        kai1 += sum(PROB[(i-1) * 5000:(i-1) * 5000 + 2000])
    test=[]
    nun=0
    for thread in thread_list:
        thread.join()
        test.append(thread.get_value())
        nun += 1
        print(nun)
    kai = [0.0] * 92
    for i in test:
        for j in range(92):
            kai[j] += i[j]
    book = openpyxl.load_workbook("test_30.xlsx")
    sheet = book[str(args.name_kei)]
    for jj in range(92):
        sheet.cell(row=2+jj,column=1).value = float(kai[jj])
    sheet.cell(row=1, column=1).value = float(kai1)
    book.save('test_30.xlsx')

if __name__ == '__main__':
    main()
