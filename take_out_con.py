
import tool
import numpy as np
import MDAnalysis
from argparse import ArgumentParser
import openpyxl
a=[]

def get_option():
    argparser = ArgumentParser()
    argparser.add_argument("-trr","--trajectory",type=str,help="path of trr")
    argparser.add_argument("-pdb","--protein",type=str,help="path of pdb")
    argparser.add_argument("-prob","--probtxt",type=str,help="path of prob")
    argparser.add_argument("-name","--name_kei",type=str,help="name of file")
    return argparser.parse_args()

def CAL_LI(cor,num):
    num1 = 0
    global MED
    for j in range(92,114):
        for j1 in MED[j]:
            a = tool.contact(cor[num][0],cor[num][1],cor[num][2],cor[j1][0],cor[j1][1],cor[j1][2])
            if a <= 8:
                num1 += 1
                break
            elif a >=15:
                break
        if num1 == 1:
            break
    return num1


def CAL_COR_MED(cor):
    global MED
    kai = []
    for i in range(92):
        for j1 in MED[i]:
            jj = CAL_LI(cor,j1)
            if jj == 1:
                break
        kai.append(int(jj))
    return kai

def CAL_TRR(prob,args):
    global MED
    u = MDAnalysis.Universe(str(args.trajectory))
    num2 = len(u.trajectory)/10
    frm = iter(u.trajectory)
    del u
    MED26 = [0.0 for i in range(92)]
    num1 = 0
    ROW = 1
    book = openpyxl.load_workbook(filename='test.xlsx')
    book.create_sheet(str(args.name_kei))
    book.save("test.xlsx")
    book.close()
    nnnn = 10 ** -100
    for i in prob:
        num1 += 1
        print(num1,"/",int(num2)*10,"  ",i)
        if i > nnnn:
            cor = next(frm)
            aa = CAL_COR_MED(cor)
            for jj in range(len(MED26)):
                if aa[jj] == 1:
                    MED26[jj] += i
        else:
            next(frm)

        if int(num1) % int(num2) == 0:
            book = openpyxl.load_workbook("test.xlsx")
            sheet = book[str(args.name_kei)]
            sheet.cell(row=1,column=int(ROW)).value=num1
            for jj in range(92):
                sheet.cell(row=2+jj,column=int(ROW)).value=float(MED26[jj])
            MED26 = [0.0 for i in range(92)]
            ROW += 1
            book.save('test.xlsx')



def main():
    global a
    args = get_option()
    a.append(str(args.trajectory).split(","))
    a.append(str(args.protein).split(","))
    a.append(str(args.probtxt).split(","))
    prob = []
    for i in open(str(args.probtxt), 'r'):
        prob.append(float(i))
    num1 = 1
    num2 = 0
    kai = []
    MED = []
    for i in open(str(args.protein)):
        f = i.split()
        if int(f[5]) == num1:
            kai.append(int(f[1]))
        else:
            MED.append(kai)
            kai=[int(f[1])]
            num1 += 1
            if f[4] == "B" and num2 == 0:
                num2 += 1
                num1 = 1
    MED.append(kai)
    global MED
    CAL_TRR(prob,args)

if __name__ == '__main__':
    main()
